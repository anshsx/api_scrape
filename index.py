from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.responses import JSONResponse
from docx import Document
import openpyxl
from pdfminer.high_level import extract_text as extract_pdf_text
import os

app = FastAPI()

def extract_text_from_file(file_path: str, file_type: str) -> str:
    try:
        if file_type in ["doc", "docx"]:
            return extract_text_from_docx(file_path)
        elif file_type in ["xls", "xlsx"]:
            return extract_text_from_excel(file_path)
        elif file_type == "pdf":
            return extract_text_from_pdf(file_path)
        elif file_type == "txt":
            return extract_text_from_txt(file_path)
        else:
            raise ValueError("Unsupported file type")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting text: {str(e)}")

def extract_text_from_docx(file_path: str) -> str:
    doc = Document(file_path)
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])

def extract_text_from_excel(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    text = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            text.append(" ".join([str(cell.value) if cell.value else "" for cell in row]))
    return "\n".join(text)

def extract_text_from_pdf(file_path: str) -> str:
    return extract_pdf_text(file_path)

def extract_text_from_txt(file_path: str) -> str:
    with open(file_path, 'r') as f:
        return f.read()

@app.post("/extract-text/")
async def extract_text(file: UploadFile):
    file_location = f"/tmp/{file.filename}"
    with open(file_location, "wb") as f:
        f.write(await file.read())

    file_type = file.filename.split(".")[-1].lower()
    text = extract_text_from_file(file_location, file_type)

    # Clean up
    os.remove(file_location)

    return JSONResponse(content={"filename": file.filename, "content": text})
